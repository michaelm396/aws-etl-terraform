from __future__ import annotations
"""Transform Lambda for the ETL pipeline.

This stage reads the raw workbook from S3, normalizes fields, converts the
categorical gender column to a gender_code integer, enriches public IP
addresses with offline geolocation metadata, and writes a processed CSV back
to S3.
"""

import csv
import ipaddress
import os
from io import BytesIO, StringIO
from pathlib import PurePosixPath
from urllib.parse import unquote_plus

import boto3
from geolite2 import geolite2
from openpyxl import load_workbook


GENDER_MAP = {
    "Male": 0,
    "Female": 1,
    "Non-binary": 2,
    "Bigender": 3,
    "Genderfluid": 4,
    "Agender": 5,
    "Genderqueer": 6,
    "Polygender": 7,
}

s3_client = boto3.client("s3")
RAW_SAMPLE_ROWS = 3
PROCESSED_SAMPLE_ROWS = 3
GEOLOCATION_HEADERS = [
    "country",
    "region",
    "city",
    "latitude",
    "longitude",
    "timezone",
]
geoip_reader = geolite2.reader()


def normalize_value(value: object) -> object:
    """Trim string values and convert blank strings to nulls."""
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    return value


def transform_gender(value: object) -> object:
    """Map the normalized gender label to its integer code."""
    if value is None:
        return value

    if value not in GENDER_MAP:
        raise ValueError(f"Unsupported gender value: {value}")

    return GENDER_MAP[value]


def null_geolocation_fields() -> list[object]:
    """Return an empty geolocation payload for failed lookups."""
    return [None] * len(GEOLOCATION_HEADERS)


def validate_public_ip(ip_value: object) -> object | None:
    """Return a parsed public IP address or None for missing/private/local values."""
    if ip_value is None:
        return None

    try:
        parsed_ip = ipaddress.ip_address(str(ip_value))
    except ValueError:
        print(f"Geolocation skipped for invalid IP address: {ip_value}")
        return None

    if (
        parsed_ip.is_private
        or parsed_ip.is_loopback
        or parsed_ip.is_multicast
        or parsed_ip.is_reserved
        or parsed_ip.is_link_local
        or parsed_ip.is_unspecified
    ):
        print(f"Geolocation skipped for non-public IP address: {ip_value}")
        return None

    return parsed_ip


def safe_nested_get(value: dict | list | None, *keys: object) -> object | None:
    """Safely walk nested dict/list values from the GeoLite2 lookup payload."""
    current: object | None = value
    for key in keys:
        if current is None:
            return None
        if isinstance(current, dict):
            current = current.get(key)
        elif isinstance(current, list) and isinstance(key, int):
            if key >= len(current):
                return None
            current = current[key]
        else:
            return None
    return current


def geolocate_ip_address(ip_value: object) -> list[object]:
    """Return geolocation fields for a public IP, or nulls when lookup fails."""
    parsed_ip = validate_public_ip(ip_value)
    if parsed_ip is None:
        return null_geolocation_fields()

    try:
        geo_record = geoip_reader.get(str(parsed_ip))
    except Exception as exc:
        print(f"Geolocation lookup failed for IP {ip_value}: {exc}")
        return null_geolocation_fields()

    if not geo_record:
        print(f"Geolocation lookup returned no data for IP {ip_value}")
        return null_geolocation_fields()

    return [
        safe_nested_get(geo_record, "country", "names", "en"),
        safe_nested_get(geo_record, "subdivisions", 0, "names", "en"),
        safe_nested_get(geo_record, "city", "names", "en"),
        safe_nested_get(geo_record, "location", "latitude"),
        safe_nested_get(geo_record, "location", "longitude"),
        safe_nested_get(geo_record, "location", "time_zone"),
    ]


def destination_key(source_key: str, processed_prefix: str) -> str:
    """Build the processed CSV key from the raw workbook key."""
    source_path = PurePosixPath(source_key)
    stem = source_path.stem
    return f"{processed_prefix.strip('/')}/{stem}_transformed.csv"


def log_sample_rows(label: str, headers: list[str], rows: list[list[object]]) -> None:
    """Log a small preview of the data flowing through the transform stage."""
    print(f"{label} sample ({len(rows)} row(s)):")
    print(headers)
    for row in rows:
        print(row)


def extract_workbook_rows_from_s3(bucket_name: str, source_key: str) -> list[tuple[object, ...]]:
    """Extract workbook rows from the raw S3 object."""
    response = s3_client.get_object(Bucket=bucket_name, Key=source_key)
    workbook = load_workbook(
        filename=BytesIO(response["Body"].read()),
        read_only=True,
        data_only=True,
    )
    worksheet = workbook.active
    return list(worksheet.iter_rows(values_only=True))


def transform_workbook_rows_to_csv(
    rows: list[tuple[object, ...]],
) -> tuple[
    str,
    list[str],
    list[str],
    list[list[object]],
    list[list[object]],
]:
    """Transform raw workbook rows into a processed CSV payload."""
    if not rows:
        raise ValueError("Uploaded workbook is empty.")

    raw_headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    processed_headers = raw_headers.copy()
    try:
        gender_index = next(
            index
            for index, header in enumerate(processed_headers)
            if header.lower() == "gender"
        )
    except StopIteration as exc:
        raise ValueError("Workbook does not contain a 'gender' column.") from exc

    try:
        ip_address_index = next(
            index
            for index, header in enumerate(processed_headers)
            if header.lower() == "ip_address"
        )
    except StopIteration as exc:
        raise ValueError("Workbook does not contain an 'ip_address' column.") from exc

    processed_headers[gender_index] = "gender_code"
    processed_headers.extend(GEOLOCATION_HEADERS)

    # Capture a tiny before/after sample for CloudWatch verification.
    raw_sample_rows = [
        [normalize_value(value) for value in row]
        for row in rows[1 : 1 + RAW_SAMPLE_ROWS]
    ]

    output_buffer = StringIO()
    writer = csv.writer(output_buffer)
    writer.writerow(processed_headers)
    processed_sample_rows: list[list[object]] = []

    for row in rows[1:]:
        mutable_row = [normalize_value(value) for value in row]
        mutable_row[gender_index] = transform_gender(mutable_row[gender_index])
        mutable_row.extend(geolocate_ip_address(mutable_row[ip_address_index]))
        writer.writerow(mutable_row)
        if len(processed_sample_rows) < PROCESSED_SAMPLE_ROWS:
            processed_sample_rows.append(mutable_row.copy())

    return (
        output_buffer.getvalue(),
        raw_headers,
        processed_headers,
        raw_sample_rows,
        processed_sample_rows,
    )


def load_transformed_csv_to_s3(
    bucket_name: str,
    transformed_key: str,
    csv_payload: str,
) -> None:
    """Load the processed CSV payload back into S3."""
    s3_client.put_object(
        Bucket=bucket_name,
        Key=transformed_key,
        Body=csv_payload.encode("utf-8"),
        ContentType="text/csv",
    )


def lambda_handler(event: dict, context: object) -> dict:
    """Handle S3-created events for raw workbook uploads."""
    processed_prefix = os.environ.get("PROCESSED_PREFIX", "processed")
    raw_prefix = os.environ.get("RAW_PREFIX", "raw").strip("/")

    for record in event.get("Records", []):
        bucket_name = record["s3"]["bucket"]["name"]
        source_key = unquote_plus(record["s3"]["object"]["key"])
        expected_prefix = f"{raw_prefix}/"

        # Ignore uploads outside the raw prefix so the function stays focused on
        # the extract-to-transform handoff.
        if not source_key.startswith(expected_prefix):
            continue

        print(f"Processing workbook from s3://{bucket_name}/{source_key}")

        rows = extract_workbook_rows_from_s3(bucket_name, source_key)
        (
            csv_payload,
            raw_headers,
            processed_headers,
            raw_sample_rows,
            processed_sample_rows,
        ) = transform_workbook_rows_to_csv(rows)
        log_sample_rows("Raw workbook", raw_headers, raw_sample_rows)
        log_sample_rows("Processed CSV", processed_headers, processed_sample_rows)

        transformed_key = destination_key(source_key, processed_prefix)
        load_transformed_csv_to_s3(bucket_name, transformed_key, csv_payload)

        print(f"Wrote transformed output to s3://{bucket_name}/{transformed_key}")

    return {
        "statusCode": 200,
        "rawPrefix": raw_prefix,
        "processedPrefix": processed_prefix,
    }
